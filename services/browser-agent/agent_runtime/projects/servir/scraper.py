"""
SERVIR Ofertas Laborales scraper.
Ruta destino: services/browser-agent/agent_runtime/projects/servir/scraper.py

Correcciones aplicadas vs versión anterior
==========================================
1. _go_to_next_page: 3 reintentos, timeouts 60 s / 30 s, espera extra 1 500 ms
   tras cambio de etiqueta (race condition PrimeFaces).
2. _parse_remuneracion: "S/. 4,000.00" → float 4000.0
3. _parse_fecha: "DD/MM/YYYY" → datetime.date (Excel escribe fecha real, no texto)
4. _parse_vacantes: "1" → int
5. _write_excel_formatted / _write_excel: tipos nativos con formatos de celda correctos.
6. NUEVO _extract_detail_page: extrae Requerimiento, Experiencia, Formación Académica,
   Especialización desde la página de detalle. Se activa con payload get_detail=True.
   Se detiene antes de "Conocimientos". WARNING: ~3-4 horas extra para 3 200 ofertas.

Mapa de columnas Excel
======================
Col  1 → Título de Convocatoria
Col  2 → Entidad
Col  3 → Ubicación
Col  4 → Número de Convocatoria
Col  5 → Vacantes            (int)
Col  6 → Remuneración        (float, #,##0.00)
Col  7 → Fecha Inicio        (date, DD/MM/YYYY)
Col  8 → Fecha Fin           (date, DD/MM/YYYY)
Col  9 → Requerimiento       (texto — sólo si get_detail=True)
Col 10 → Experiencia         (texto — sólo si get_detail=True)
Col 11 → Formación Académica (texto — sólo si get_detail=True)
Col 12 → Especialización     (texto — sólo si get_detail=True)
Col 13 → No me interesa      (marcado manual)

IMPORTANTE sobre los links de convocatoria
==========================================
SERVIR usa JSF puro (PrimeFaces). "¡Ver más!" hace un POST sin parámetros en la URL.
No existe URL estable por oferta. El N° de Aviso sólo aparece en la página de detalle.
"""

import json
import logging
import os
import re
from datetime import date, datetime, timezone
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select

from ...browser_manager import BrowserManager
from ...database import SessionLocal
from .models import ServirOferta

logger = logging.getLogger("agent_runtime.scrapers.servir")

# ---------------------------------------------------------------------------
# Progreso en Redis (para el dashboard de monitoreo)
# ---------------------------------------------------------------------------
_REDIS_PROGRESS_KEY = "servir:progress"
_REDIS_TTL_SECONDS  = 86_400   # 24 h


def _get_redis():
    """Devuelve un cliente Redis sin fallar si Redis no está disponible."""
    try:
        import redis as _redis_lib
        url = os.environ.get("REDIS_URL", "redis://redis:6379")
        return _redis_lib.from_url(url, socket_connect_timeout=2)
    except Exception:
        return None


def _report_progress(
    status: str,
    current_page: int,
    total_pages: int,
    offers_scraped: int,
    started_at: str,
    error: str | None = None,
) -> None:
    """Escribe el progreso del scraping en Redis para que el dashboard lo lea."""
    payload = {
        "status":         status,          # "running" | "done" | "error"
        "current_page":   current_page,
        "total_pages":    total_pages,
        "offers_scraped": offers_scraped,
        "started_at":     started_at,
        "updated_at":     datetime.now(timezone.utc).isoformat(),
        "error":          error,
    }
    try:
        r = _get_redis()
        if r:
            r.set(_REDIS_PROGRESS_KEY, json.dumps(payload), ex=_REDIS_TTL_SECONDS)
    except Exception as exc:
        logger.debug("No se pudo actualizar progreso en Redis: %s", exc)

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------
DEFAULT_URL          = "https://app.servir.gob.pe/DifusionOfertasExterno/faces/consultas/ofertas_laborales.xhtml"
CARDS_SELECTOR       = "form#frmLstOfertsLabo .cuadro-vacantes"
PAGE_LABEL_SELECTOR  = "label.btn-paginator-cnt"
NEXT_BUTTON_SELECTOR = "button:has-text('Sig.')"
VER_MAS_SELECTOR     = "a:has-text('¡Ver más!'), button:has-text('¡Ver más!')"

MARK_COLUMN_HEADER = "No me interesa"
MARK_VALUES        = {"X", "SI", "SÍ", "YES", "1", "S"}

EXCEL_HEADERS = [
    "Título de Convocatoria",    # col  1
    "Entidad",                   # col  2
    "Ubicación",                 # col  3
    "Número de Convocatoria",    # col  4
    "Vacantes",                  # col  5  → int
    "Remuneración",              # col  6  → float "#,##0.00"
    "Fecha Inicio Publicación",  # col  7  → date  "DD/MM/YYYY"
    "Fecha Fin Publicación",     # col  8  → date  "DD/MM/YYYY"
    "Requerimiento",             # col  9  → texto (detalle)
    "Experiencia",               # col 10  → texto (detalle)
    "Formación Académica",       # col 11  → texto (detalle)
    "Especialización",           # col 12  → texto (detalle)
    MARK_COLUMN_HEADER,          # col 13  → marcado manual
]

# ---------------------------------------------------------------------------
# JS — extrae tarjetas del listado
# ---------------------------------------------------------------------------
EXTRACT_CARDS_JS = """
() => {
  const cards = Array.from(
    document.querySelectorAll('form#frmLstOfertsLabo .cuadro-vacantes')
  );
  return cards.map(card => {
    const titulo  = card.querySelector('.titulo-vacante label')
                        ?.innerText?.trim() || '';
    const entidad = card.querySelector('.nombre-entidad .detalle-sp')
                        ?.innerText?.trim() || '';
    const fields  = {};
    card.querySelectorAll('.col-sm-5 .row.box-mb .col-sm-12').forEach(row => {
      const label = row.querySelector('.sub-titulo')
                        ?.innerText?.trim().replace(/:\\s*$/, '') || '';
      const value = row.querySelector('.detalle-sp')
                        ?.innerText?.trim() || '';
      if (label) fields[label] = value;
    });
    return {
      titulo,
      entidad,
      ubicacion:            fields['Ubicación']                   || '',
      numero_convocatoria:  fields['Número de Convocatoria']      || '',
      vacantes:             fields['Cantidad de Vacantes']        || '',
      remuneracion:         fields['Remuneración']                || '',
      fecha_inicio:         fields['Fecha Inicio de Publicación'] || '',
      fecha_fin:            fields['Fecha Fin de Publicación']    || '',
    };
  });
}
"""

# ---------------------------------------------------------------------------
# JS — extrae campos de la página de DETALLE
# Captura: aviso_num, requerimiento, experiencia, formacion_academica,
#          especializacion. Se detiene al encontrar "Conocimientos".
# ---------------------------------------------------------------------------
EXTRACT_DETAIL_JS = r"""
() => {
  const STOP_KEY = /conocimientos/i;
  const TARGET_KEYS = {
    requerimiento:       /^requerimiento/i,
    experiencia:         /^experiencia/i,
    formacion_academica: /^formaci[oó]n\s+acad[eé]mica/i,
    especializacion:     /^especializaci[oó]n/i,
  };

  function nodeText(el) {
    return (el.innerText || el.textContent || '').trim();
  }

  const result = {
    aviso_num:           '',
    requerimiento:       '',
    experiencia:         '',
    formacion_academica: '',
    especializacion:     '',
  };

  const bodyText = document.body.innerText || '';
  const avisoMatch = bodyText.match(/N[°º]\s*(\d{4,8})/);
  if (avisoMatch) result.aviso_num = avisoMatch[1];

  const allEls  = Array.from(document.body.querySelectorAll('*'));
  let currentKey = null;
  let stop = false;

  for (const el of allEls) {
    if (stop) break;
    const ownText = nodeText(el);
    if (!ownText) continue;
    if (STOP_KEY.test(ownText)) { stop = true; break; }

    let matchedKey = null;
    for (const [key, pattern] of Object.entries(TARGET_KEYS)) {
      if (pattern.test(ownText.replace(/:\s*$/, ''))) {
        matchedKey = key; break;
      }
    }

    if (matchedKey) {
      currentKey = matchedKey;
      const colonIdx = ownText.indexOf(':');
      if (colonIdx !== -1) {
        const inline = ownText.slice(colonIdx + 1).trim();
        if (inline) { result[currentKey] = inline; currentKey = null; }
      }
      continue;
    }

    if (currentKey && ownText.length > 1 && !result[currentKey]) {
      result[currentKey] = ownText;
      currentKey = null;
    }
  }

  for (const k of Object.keys(result)) {
    if (typeof result[k] === 'string') {
      result[k] = result[k].replace(/\s{2,}/g, ' ').trim();
    }
  }
  return result;
}
"""

# ---------------------------------------------------------------------------
# Helpers de parseo
# ---------------------------------------------------------------------------

def _parse_remuneracion(raw: str) -> float | None:
    """"S/. 4,000.00" → 4000.0"""
    if not raw:
        return None
    cleaned = re.sub(r"[Ss]/\.\s*", "", raw).strip().replace(",", "")
    try:
        return float(cleaned)
    except ValueError:
        logger.warning("No se pudo parsear remuneracion: %r", raw)
        return None


def _parse_fecha(raw: str) -> date | None:
    """"DD/MM/YYYY" → datetime.date"""
    if not raw:
        return None
    try:
        return datetime.strptime(raw.strip(), "%d/%m/%Y").date()
    except ValueError:
        logger.warning("No se pudo parsear fecha: %r", raw)
        return None


def _parse_vacantes(raw: str) -> int | None:
    """"1" → 1"""
    if not raw:
        return None
    try:
        return int(raw.strip())
    except ValueError:
        logger.warning("No se pudo parsear vacantes: %r", raw)
        return None


def _empty_detail() -> dict:
    return {
        "aviso_num": "", "requerimiento": "", "experiencia": "",
        "formacion_academica": "", "especializacion": "",
    }


# ---------------------------------------------------------------------------
# Paginador
# ---------------------------------------------------------------------------

def _get_page_label_text(page) -> str:
    try:
        return page.locator(PAGE_LABEL_SELECTOR).first.inner_text(timeout=5_000).strip()
    except Exception:
        return ""


def _parse_total_pages(label_text: str) -> int:
    match = re.search(r"P[aá]gina\s+\d+\s+de\s+(\d+)", label_text)
    if not match:
        raise ValueError(f"No se pudo interpretar el texto de paginación: '{label_text}'")
    return int(match.group(1))


def _humanized_pause(page, page_num: int = 0) -> None:
    import random
    if page_num and page_num % 20 == 0:
        delay = random.uniform(8, 15)
    else:
        delay = random.uniform(2.5, 5.5)
    page.wait_for_timeout(int(delay * 1000))


def _go_to_next_page(page, current_label: str) -> bool:
    """
    Hace clic en 'Sig.' y espera a que PrimeFaces actualice la página.
    Correcciones:
    - 3 reintentos
    - timeout POST: 60 000 ms (antes 30 000)
    - timeout label: 30 000 ms (antes 15 000)
    - espera extra 1 500 ms tras cambio de etiqueta (race condition PrimeFaces)
    Devuelve True si tuvo éxito, False si falló tras 3 intentos.
    """
    next_btn = page.locator(NEXT_BUTTON_SELECTOR).first
    try:
        if next_btn.is_disabled(timeout=3_000):
            return False
    except Exception:
        return False

    for attempt in range(1, 4):
        try:
            with page.expect_response(
                lambda r: "ofertas_laborales.xhtml" in r.url and r.request.method == "POST",
                timeout=60_000,
            ):
                next_btn.click()

            page.wait_for_function(
                """(oldLabel) => {
                    const el = document.querySelector('label.btn-paginator-cnt');
                    return el && el.innerText.trim() !== oldLabel;
                }""",
                arg=current_label,
                timeout=30_000,
            )
            # Espera extra: PrimeFaces cambia el label ANTES de renderizar tarjetas
            page.wait_for_timeout(1_500)
            return True

        except Exception as exc:
            logger.warning("Intento %d/3 de pasar página falló: %s", attempt, exc)
            if attempt < 3:
                page.wait_for_timeout(3_000)

    logger.error("No se pudo pasar de página tras 3 intentos.")
    return False


# ---------------------------------------------------------------------------
# Extracción de página de detalle
# ---------------------------------------------------------------------------

def _extract_detail_page(page, context, card_index: int) -> dict:
    """
    Hace clic en '¡Ver más!' de la tarjeta `card_index` (base-0).
    Extrae aviso_num, requerimiento, experiencia, formacion_academica,
    especializacion. Vuelve al listado navegando a DEFAULT_URL (no go_back),
    para obtener un ViewState JSF fresco y evitar corromper el estado del paginador.
    Devuelve dict con valores vacíos si falla.
    """
    navigated = False
    detail     = _empty_detail()

    try:
        cards = page.locator(CARDS_SELECTOR).all()
        if card_index >= len(cards):
            return _empty_detail()

        ver_mas = cards[card_index].locator(VER_MAS_SELECTOR).first
        with page.expect_response(
            lambda r: "detalle_ofertas_laborales.xhtml" in r.url
                      and r.request.method == "POST",
            timeout=30_000,
        ):
            ver_mas.click()

        navigated = True
        page.wait_for_load_state("networkidle", timeout=20_000)
        detail = page.evaluate(EXTRACT_DETAIL_JS)

    except Exception as exc:
        logger.warning("Error al abrir detalle (tarjeta %d): %s", card_index, exc)

    if navigated:
        try:
            # Navegar a DEFAULT_URL en lugar de go_back() para obtener un
            # ViewState JSF fresco. El servidor mantiene la posición de página
            # en la sesión HTTP, por lo que debería retornar a la misma página.
            page.goto(DEFAULT_URL, wait_until="networkidle", timeout=30_000)
            # Esperar que PrimeFaces inicialice los botones "Ver más"
            # antes de intentar el siguiente click.
            page.wait_for_selector(VER_MAS_SELECTOR, state="visible", timeout=15_000)
            page.wait_for_timeout(1_000)
        except Exception as exc:
            logger.warning("Error al restaurar listado tras detalle (tarjeta %d): %s", card_index, exc)

    return detail


# ---------------------------------------------------------------------------
# Estilos Excel
# ---------------------------------------------------------------------------

DATE_FMT     = "DD/MM/YYYY"
CURR_FMT     = "#,##0.00"
HEADER_FILL  = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT  = Font(color="FFFFFF", bold=True)
ROW_FILL_ALT = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=True)

COLUMN_WIDTHS = [45, 38, 26, 24, 10, 16, 16, 16, 50, 50, 50, 50, 16]


def _write_excel_formatted(rows: list[dict], output_path: str, dup_rows: list[dict] | None = None) -> None:
    """Escribe Excel con tipos correctos: int, float, date y texto."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ofertas SERVIR"

    # Cabeceras
    ws.append([h or "" for h in EXCEL_HEADERS])
    for col_idx in range(1, len(EXCEL_HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Datos
    for row_num, row in enumerate(rows, start=2):
        alt = (row_num % 2 == 0)

        def _put(col: int, value, fmt: str | None = None, align=LEFT):
            c = ws.cell(row=row_num, column=col, value=value)
            if alt:
                c.fill = ROW_FILL_ALT
            c.alignment = align
            if fmt:
                c.number_format = fmt
            return c

        _put(1,  row.get("titulo",              ""))
        _put(2,  row.get("entidad",             ""))
        _put(3,  row.get("ubicacion",           ""))
        _put(4,  row.get("numero_convocatoria", ""), align=CENTER)
        _put(5,  _parse_vacantes(row.get("vacantes",      "")))
        _put(6,  _parse_remuneracion(row.get("remuneracion", "")), CURR_FMT, CENTER)
        _put(7,  _parse_fecha(row.get("fecha_inicio", "")),        DATE_FMT, CENTER)
        _put(8,  _parse_fecha(row.get("fecha_fin",    "")),        DATE_FMT, CENTER)
        _put(9,  row.get("requerimiento",       ""))
        _put(10, row.get("experiencia",         ""))
        _put(11, row.get("formacion_academica", ""))
        _put(12, row.get("especializacion",     ""))
        _put(13, None)

    # Anchos
    for col_idx, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Hoja Duplicados ─────────────────────────────────────────────────
    if dup_rows:
        ws_dup = wb.create_sheet(title="Duplicados")
        ws_dup.append([h or "" for h in EXCEL_HEADERS])
        DUP_FILL = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
        for col_idx in range(1, len(EXCEL_HEADERS) + 1):
            cell = ws_dup.cell(row=1, column=col_idx)
            cell.fill = DUP_FILL
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = CENTER
        ws_dup.row_dimensions[1].height = 30
        ws_dup.freeze_panes = "A2"
        for row_num_d, row_d in enumerate(dup_rows, start=2):
            alt_d = (row_num_d % 2 == 0)
            def _put_d(col, value, fmt=None, align=LEFT, _rn=row_num_d, _alt=alt_d):
                c = ws_dup.cell(row=_rn, column=col, value=value)
                if _alt:
                    c.fill = ROW_FILL_ALT
                c.alignment = align
                if fmt:
                    c.number_format = fmt
                return c
            _put_d(1,  row_d.get("titulo",              ""))
            _put_d(2,  row_d.get("entidad",             ""))
            _put_d(3,  row_d.get("ubicacion",           ""))
            _put_d(4,  row_d.get("numero_convocatoria", ""), align=CENTER)
            _put_d(5,  _parse_vacantes(row_d.get("vacantes",      "")))
            _put_d(6,  _parse_remuneracion(row_d.get("remuneracion", "")), CURR_FMT, CENTER)
            _put_d(7,  _parse_fecha(row_d.get("fecha_inicio", "")),        DATE_FMT, CENTER)
            _put_d(8,  _parse_fecha(row_d.get("fecha_fin",    "")),        DATE_FMT, CENTER)
            _put_d(9,  row_d.get("requerimiento",       ""))
            _put_d(10, row_d.get("experiencia",         ""))
            _put_d(11, row_d.get("formacion_academica", ""))
            _put_d(12, row_d.get("especializacion",     ""))
            _put_d(13, None)
        for col_idx, width in enumerate(COLUMN_WIDTHS, start=1):
            ws_dup.column_dimensions[get_column_letter(col_idx)].width = width
    wb.save(output_path)
    logger.info("Excel guardado: %s (%d filas)", output_path, len(rows))


def _upsert_excel(
    scraped_rows: list[dict],
    output_path: str,
    dup_rows: list[dict] | None = None,
) -> tuple[int, int]:
    """
    Agrega al Excel existente únicamente las filas cuyo numero_convocatoria
    no está todavía en la hoja correspondiente.

    - Si el archivo no existe → lo crea desde cero con todos los scraped_rows.
    - Si el archivo existe    → lee los numero_convocatoria presentes, solo
                                 agrega los que son genuinamente nuevos.
    - Las marcas del usuario (col 13) en el archivo existente se conservan.
    - Misma lógica para la hoja Duplicados.

    Retorna (nuevas_main, nuevas_dup).
    """
    NUM_COL_IDX = 4   # col 4 = Número de Convocatoria (base-1, índice base-0 = 3)

    existing_main_nums: set[str] = set()
    existing_dup_nums:  set[str] = set()
    wb_exists = os.path.exists(output_path)
    wb = None

    if wb_exists:
        try:
            wb = load_workbook(output_path)
        except Exception as exc:
            logger.warning(
                "No se pudo cargar el Excel existente (%s); se recreará desde cero.", exc
            )
            wb_exists = False

    if wb_exists and wb is not None:
        ws_main = wb.active
        for xrow in ws_main.iter_rows(min_row=2, values_only=True):
            num = str(xrow[NUM_COL_IDX - 1] or "").strip()
            if num:
                existing_main_nums.add(num)

        ws_dup = wb["Duplicados"] if "Duplicados" in wb.sheetnames else None
        if ws_dup is not None:
            for xrow in ws_dup.iter_rows(min_row=2, values_only=True):
                num = str(xrow[NUM_COL_IDX - 1] or "").strip()
                if num:
                    existing_dup_nums.add(num)
    else:
        # Crear libro nuevo con cabecera
        wb = Workbook()
        ws_main = wb.active
        ws_main.title = "Ofertas SERVIR"
        ws_main.append([h or "" for h in EXCEL_HEADERS])
        for col_idx in range(1, len(EXCEL_HEADERS) + 1):
            cell = ws_main.cell(row=1, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
        ws_main.row_dimensions[1].height = 30
        ws_main.freeze_panes = "A2"
        for col_idx, width in enumerate(COLUMN_WIDTHS, start=1):
            ws_main.column_dimensions[get_column_letter(col_idx)].width = width
        ws_dup = None

    # ── Filtrar filas nuevas para hoja principal ──────────────────────────────
    new_main_rows = [
        r for r in scraped_rows
        if str(r.get("numero_convocatoria", "")).strip() not in existing_main_nums
    ]

    for row in new_main_rows:
        row_num = ws_main.max_row + 1
        alt     = (row_num % 2 == 0)

        def _put(col, value, fmt=None, align=LEFT, _rn=row_num, _alt=alt):
            c = ws_main.cell(row=_rn, column=col, value=value)
            if _alt:
                c.fill = ROW_FILL_ALT
            c.alignment = align
            if fmt:
                c.number_format = fmt
            return c

        _put(1,  row.get("titulo",              ""))
        _put(2,  row.get("entidad",             ""))
        _put(3,  row.get("ubicacion",           ""))
        _put(4,  row.get("numero_convocatoria", ""), align=CENTER)
        _put(5,  _parse_vacantes(row.get("vacantes",      "")))
        _put(6,  _parse_remuneracion(row.get("remuneracion", "")), CURR_FMT, CENTER)
        _put(7,  _parse_fecha(row.get("fecha_inicio", "")),        DATE_FMT, CENTER)
        _put(8,  _parse_fecha(row.get("fecha_fin",    "")),        DATE_FMT, CENTER)
        _put(9,  row.get("requerimiento",       ""))
        _put(10, row.get("experiencia",         ""))
        _put(11, row.get("formacion_academica", ""))
        _put(12, row.get("especializacion",     ""))
        _put(13, None)   # columna de marcado — el usuario la llena manualmente

    # ── Hoja Duplicados ───────────────────────────────────────────────────────
    nuevas_dup = 0
    if dup_rows:
        new_dup_rows = [
            r for r in dup_rows
            if str(r.get("numero_convocatoria", "")).strip() not in existing_dup_nums
        ]
        if new_dup_rows:
            if ws_dup is None:
                ws_dup = wb.create_sheet(title="Duplicados")
                ws_dup.append([h or "" for h in EXCEL_HEADERS])
                DUP_FILL = PatternFill(
                    start_color="8B0000", end_color="8B0000", fill_type="solid"
                )
                for col_idx in range(1, len(EXCEL_HEADERS) + 1):
                    cell = ws_dup.cell(row=1, column=col_idx)
                    cell.fill = DUP_FILL
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = CENTER
                ws_dup.row_dimensions[1].height = 30
                ws_dup.freeze_panes = "A2"
                for col_idx, width in enumerate(COLUMN_WIDTHS, start=1):
                    ws_dup.column_dimensions[get_column_letter(col_idx)].width = width

            for row_d in new_dup_rows:
                row_num_d = ws_dup.max_row + 1
                alt_d     = (row_num_d % 2 == 0)

                def _put_d(col, value, fmt=None, align=LEFT, _rn=row_num_d, _alt=alt_d):
                    c = ws_dup.cell(row=_rn, column=col, value=value)
                    if _alt:
                        c.fill = ROW_FILL_ALT
                    c.alignment = align
                    if fmt:
                        c.number_format = fmt
                    return c

                _put_d(1,  row_d.get("titulo",              ""))
                _put_d(2,  row_d.get("entidad",             ""))
                _put_d(3,  row_d.get("ubicacion",           ""))
                _put_d(4,  row_d.get("numero_convocatoria", ""), align=CENTER)
                _put_d(5,  _parse_vacantes(row_d.get("vacantes",      "")))
                _put_d(6,  _parse_remuneracion(row_d.get("remuneracion", "")), CURR_FMT, CENTER)
                _put_d(7,  _parse_fecha(row_d.get("fecha_inicio", "")),        DATE_FMT, CENTER)
                _put_d(8,  _parse_fecha(row_d.get("fecha_fin",    "")),        DATE_FMT, CENTER)
                _put_d(9,  row_d.get("requerimiento",       ""))
                _put_d(10, row_d.get("experiencia",         ""))
                _put_d(11, row_d.get("formacion_academica", ""))
                _put_d(12, row_d.get("especializacion",     ""))
                _put_d(13, None)
            nuevas_dup = len(new_dup_rows)

    wb.save(output_path)
    nuevas_main = len(new_main_rows)
    logger.info(
        "Excel actualizado: %s — %d nuevas en 'Ofertas SERVIR', %d nuevas en 'Duplicados'",
        output_path, nuevas_main, nuevas_dup,
    )
    return nuevas_main, nuevas_dup


def _write_excel(rows: list[dict], output_path: str) -> None:
    """Versión minimal sin estilos, con tipos de dato correctos."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ofertas SERVIR"
    ws.append([h or "" for h in EXCEL_HEADERS])

    for row in rows:
        ws.append([
            row.get("titulo",              ""),
            row.get("entidad",             ""),
            row.get("ubicacion",           ""),
            row.get("numero_convocatoria", ""),
            _parse_vacantes(row.get("vacantes",      "")),
            _parse_remuneracion(row.get("remuneracion", "")),
            _parse_fecha(row.get("fecha_inicio", "")),
            _parse_fecha(row.get("fecha_fin",    "")),
            row.get("requerimiento",       ""),
            row.get("experiencia",         ""),
            row.get("formacion_academica", ""),
            row.get("especializacion",     ""),
            None,
        ])

    for row_cells in ws.iter_rows(min_row=2, min_col=6, max_col=8):
        row_cells[0].number_format = CURR_FMT
        row_cells[1].number_format = DATE_FMT
        row_cells[2].number_format = DATE_FMT

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Lectura de marcados ("No me interesa" en col 15 → índice 14)
# ---------------------------------------------------------------------------

def _read_marked_keys(output_path: str) -> set[tuple[str, str, str]]:
    if not os.path.exists(output_path):
        return set()
    try:
        wb = load_workbook(output_path, data_only=True)
        ws = wb.active
        header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    except Exception as exc:
        logger.warning("No se pudo leer el Excel previo: %s", exc)
        return set()

    try:
        idx_titulo = header_row.index("Título de Convocatoria")
        idx_num    = header_row.index("Número de Convocatoria")
        idx_ent    = header_row.index("Entidad")
        idx_mark   = header_row.index(MARK_COLUMN_HEADER)
    except ValueError:
        logger.warning("El Excel previo no tiene las columnas esperadas.")
        return set()

    marked: set[tuple[str, str, str]] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if idx_mark >= len(row):
            continue
        mark_value = row[idx_mark]
        if mark_value and str(mark_value).strip().upper() in MARK_VALUES:
            titulo  = str(row[idx_titulo] or "").strip()
            numero  = str(row[idx_num]    or "").strip()
            entidad = str(row[idx_ent]    or "").strip()
            if numero and entidad and titulo:
                marked.add((numero, entidad, titulo))
    return marked


# ---------------------------------------------------------------------------
# sync_servir_daily — sincronización diaria con BD + Excel
# ---------------------------------------------------------------------------

def sync_servir_daily(payload: dict[str, Any], browser: BrowserManager) -> dict[str, Any]:
    """
    Scrapea SERVIR, actualiza BD y genera Excel.

    Parámetros del payload
    ----------------------
    url          : str   URL del listado (default DEFAULT_URL)
    session_label: str   Etiqueta de sesión del navegador (default "servir")
    max_pages    : int   Límite de páginas (default None = todas)
    filename     : str   Nombre del archivo Excel de salida
    get_detail   : bool  Si True, hace clic en "¡Ver más!" por cada tarjeta
                         para extraer Requerimiento, Experiencia, Formación
                         Académica, Especialización. ~3-4 h extra. Default False.
    """
    url           = payload.get("url",           DEFAULT_URL)
    session_label = payload.get("session_label", "servir")
    max_pages     = payload.get("max_pages")
    get_detail    = bool(payload.get("get_detail", False))

    output_filename = payload.get("filename", "Ofertas_SERVIR_activas.xlsx")
    output_path     = f"/app/output/{output_filename}"

    run_time = datetime.now(timezone.utc)

    # Procesar marcados "No me interesa" del Excel previo
    marked_keys = _read_marked_keys(output_path)
    db = SessionLocal()
    try:
        for numero, entidad, titulo in marked_keys:
            oferta = db.scalar(
                select(ServirOferta).where(
                    ServirOferta.numero_convocatoria == numero,
                    ServirOferta.entidad             == entidad,
                    ServirOferta.titulo              == titulo,
                )
            )
            if oferta and not oferta.removed_by_user:
                oferta.removed_by_user    = True
                oferta.removed_by_user_at = run_time
        db.commit()
    finally:
        db.close()

    # Scraping
    context = browser.get_context(session_label)
    page    = context.new_page()

    all_rows:     list[dict] = []
    failed_pages: list[int]  = []
    pages_scraped  = 0
    stopped_early  = False
    error_detail   = None

    try:
        page.goto(url, wait_until="load", timeout=60_000)
        page.wait_for_timeout(4_000)

        label_text  = _get_page_label_text(page)
        total_pages = _parse_total_pages(label_text)
        target_pages = min(max_pages, total_pages) if max_pages else total_pages
        logger.info(
            "SERVIR (sync diario): %d páginas totales, recorriendo %d (get_detail=%s)",
            total_pages, target_pages, get_detail,
        )
        started_at_iso = run_time.isoformat()
        _report_progress("running", 0, target_pages, 0, started_at_iso)

        current_page_num = 1
        while current_page_num <= target_pages:
            try:
                cards_data: list[dict] = page.evaluate(EXTRACT_CARDS_JS)

                if get_detail:
                    for idx, card in enumerate(cards_data):
                        logger.debug(
                            "  Detalle pág %d, tarjeta %d/%d ...",
                            current_page_num, idx + 1, len(cards_data),
                        )
                        detail = _extract_detail_page(page, context, idx)
                        card.update(detail)
                        page.wait_for_timeout(700)
                        # Verificar que seguimos en la página correcta tras
                        # la navegación a DEFAULT_URL en _extract_detail_page.
                        # Si el servidor nos mandó a página 1, el paginador
                        # mostrará "1" en vez del page_num actual.
                        try:
                            lbl_check = _get_page_label_text(page)
                            if str(current_page_num) not in lbl_check:
                                logger.warning(
                                    "Posición de página perdida tras detalle "
                                    "(esperado %d, etiqueta: '%s'). Renavegando...",
                                    current_page_num, lbl_check,
                                )
                                # Re-navegar hasta current_page_num desde página 1
                                for _p in range(1, current_page_num):
                                    _lbl = _get_page_label_text(page)
                                    _go_to_next_page(page, _lbl)
                                    page.wait_for_timeout(800)
                        except Exception as _exc:
                            logger.debug("Verificación de página tras detalle: %s", _exc)
                else:
                    for card in cards_data:
                        card.update(_empty_detail())

                all_rows.extend(cards_data)
                pages_scraped += 1
            except Exception as exc:
                logger.warning("Fallo extrayendo página %d: %s", current_page_num, exc)
                failed_pages.append(current_page_num)

            # Actualizar progreso en Redis CADA página
            _report_progress("running", current_page_num, target_pages, len(all_rows), started_at_iso)

            if pages_scraped % 20 == 0 or current_page_num == target_pages:
                logger.info(
                    "SERVIR (sync diario): progreso %d/%d páginas, %d ofertas",
                    current_page_num, target_pages, len(all_rows),
                )

            if current_page_num >= target_pages:
                break

            try:
                label_text = _get_page_label_text(page)
                ok = _go_to_next_page(page, label_text)
                if not ok:
                    logger.error("Botón Sig. deshabilitado en página %d — cortando.", current_page_num)
                    stopped_early = True
                    break
            except Exception as exc:
                logger.error("Fallo avanzando desde página %d: %s", current_page_num, exc)
                stopped_early = True
                error_detail  = str(exc)
                break

            current_page_num += 1
            _humanized_pause(page, current_page_num)

    except Exception as exc:
        logger.error("Fallo irrecuperable en sync_servir_daily: %s", exc)
        stopped_early = True
        error_detail  = str(exc)
    finally:
        page.close()

    recorrido_completo = not stopped_early and not failed_pages and pages_scraped >= total_pages

    # Persistir en BD
    db = SessionLocal()
    nuevas = 0
    try:
        for row in all_rows:
            numero  = row.get("numero_convocatoria", "").strip()
            entidad = row.get("entidad",             "").strip()
            titulo  = row.get("titulo",              "").strip()
            if not numero or not entidad or not titulo:
                continue
            try:
                oferta = db.scalar(
                    select(ServirOferta).where(
                        ServirOferta.numero_convocatoria == numero,
                        ServirOferta.entidad             == entidad,
                        ServirOferta.titulo              == titulo,
                    )
                )
                if oferta:
                    oferta.ubicacion    = row.get("ubicacion",    "")
                    oferta.vacantes     = row.get("vacantes",     "")
                    oferta.remuneracion = row.get("remuneracion", "")
                    oferta.fecha_inicio = row.get("fecha_inicio", "")
                    oferta.fecha_fin    = row.get("fecha_fin",    "")
                    oferta.last_seen_at = run_time
                else:
                    db.add(ServirOferta(
                        numero_convocatoria = numero,
                        entidad             = entidad,
                        titulo              = titulo,
                        ubicacion           = row.get("ubicacion",    ""),
                        vacantes            = row.get("vacantes",     ""),
                        remuneracion        = row.get("remuneracion", ""),
                        fecha_inicio        = row.get("fecha_inicio", ""),
                        fecha_fin           = row.get("fecha_fin",    ""),
                        first_seen_at       = run_time,
                        last_seen_at        = run_time,
                        removed_by_user     = False,
                    ))
                    nuevas += 1
                db.commit()
            except Exception as exc:
                db.rollback()
                logger.warning("Fallo guardando oferta '%s': %s", titulo, exc)

        if recorrido_completo:
            activas = db.scalars(
                select(ServirOferta).where(
                    ServirOferta.last_seen_at == run_time,
                    ServirOferta.removed_by_user.is_(False),
                )
            ).all()
            vencidas = db.scalars(
                select(ServirOferta).where(ServirOferta.last_seen_at < run_time)
            ).all()
            for v in vencidas:
                db.delete(v)
            db.commit()
        else:
            activas = db.scalars(
                select(ServirOferta).where(ServirOferta.removed_by_user.is_(False))
            ).all()

        final_rows = [
            {
                "titulo":              o.titulo,
                "entidad":             o.entidad,
                "ubicacion":           o.ubicacion,
                "numero_convocatoria": o.numero_convocatoria,
                "vacantes":            o.vacantes,
                "remuneracion":        o.remuneracion,
                "fecha_inicio":        o.fecha_inicio,
                "fecha_fin":           o.fecha_fin,
                # campos de detalle vacíos en BD (sólo se guardan en Excel)
                "requerimiento":       row.get("requerimiento",       ""),
                "experiencia":         row.get("experiencia",         ""),
                "formacion_academica": row.get("formacion_academica", ""),
                "especializacion":     row.get("especializacion",     ""),
            }
            for o, row in _match_rows(activas, all_rows)
        ]
    finally:
        db.close()

    # Calcular duplicados: rows de all_rows con misma (numero, entidad, titulo)
    from collections import Counter as _Ctr
    _dup_cnt = _Ctr(
        (r.get("numero_convocatoria","").strip(),
         r.get("entidad","").strip(),
         r.get("titulo","").strip())
        for r in all_rows
    )
    _dup_rows = [
        r for r in all_rows
        if _dup_cnt[(
            r.get("numero_convocatoria","").strip(),
            r.get("entidad","").strip(),
            r.get("titulo","").strip()
        )] > 1
    ]
    # Agregar al Excel solo las filas nuevas (numero_convocatoria no presente aún)
    nuevas_excel, nuevas_dup_excel = _upsert_excel(
        all_rows, output_path, dup_rows=_dup_rows or None
    )

    result = {
        "output_file":             output_filename,
        "nuevas_en_excel":         nuevas_excel,
        "nuevas_en_duplicados":    nuevas_dup_excel,
        "nuevas_hoy":              nuevas,
        "marcadas_no_interesa":    len(marked_keys),
        "paginas_recorridas":      pages_scraped,
        "paginas_fallidas":        failed_pages,
        "recorrido_completo":      recorrido_completo,
        "detalle_error":           error_detail,
    }
    _report_progress(
        status         = "done" if recorrido_completo else "error",
        current_page   = pages_scraped,
        total_pages    = pages_scraped,   # evita mostrar "incompleto" si terminó OK
        offers_scraped = len(final_rows),
        started_at     = run_time.isoformat(),
        error          = error_detail,
    )
    logger.info("SERVIR (sync diario): finalizado — %s", result)
    return result


def _match_rows(activas, all_rows: list[dict]) -> list[tuple]:
    """
    Empareja cada oferta activa de la BD con su fila scrapeada
    (para recuperar los campos de detalle que no se guardan en BD).
    """
    index = {
        (r.get("numero_convocatoria", "").strip(),
         r.get("entidad",             "").strip(),
         r.get("titulo",              "").strip()): r
        for r in all_rows
    }
    result = []
    for o in activas:
        key = (o.numero_convocatoria.strip(), o.entidad.strip(), o.titulo.strip())
        row = index.get(key, {})
        result.append((o, row))
    return result


# ---------------------------------------------------------------------------
# scrape_servir_ofertas — extracción simple (sin sync de BD)
# ---------------------------------------------------------------------------

def scrape_servir_ofertas(payload: dict[str, Any], browser: BrowserManager) -> dict[str, Any]:
    """
    Extrae ofertas y genera Excel. No hace sync con la BD.

    Parámetros del payload
    ----------------------
    url          : str
    session_label: str   (default "servir")
    max_pages    : int   (default None = todas)
    start_page   : int   (default 1)
    filename     : str
    get_detail   : bool  (default False)
    """
    url           = payload.get("url",           DEFAULT_URL)
    session_label = payload.get("session_label", "servir")
    max_pages     = payload.get("max_pages")
    start_page    = payload.get("start_page", 1)
    get_detail    = bool(payload.get("get_detail", False))

    today    = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    filename = payload.get("filename", f"Ofertas_Laborales_SERVIR_{today}.xlsx")
    output_path = f"/app/output/{filename}"

    context = browser.get_context(session_label)
    page    = context.new_page()

    all_rows:     list[dict] = []
    failed_pages: list[int]  = []
    pages_scraped  = 0
    stopped_early  = False
    error_detail   = None

    try:
        page.goto(url, wait_until="load", timeout=60_000)
        page.wait_for_timeout(4_000)

        label_text  = _get_page_label_text(page)
        total_pages = _parse_total_pages(label_text)
        target_pages = min(max_pages, total_pages) if max_pages else total_pages
        logger.info(
            "SERVIR: %d páginas totales, recorriendo %d desde página %d (get_detail=%s)",
            total_pages, target_pages, start_page, get_detail,
        )

        # Avanzar hasta start_page
        current_page_num = 1
        while current_page_num < start_page:
            label_text = _get_page_label_text(page)
            _go_to_next_page(page, label_text)
            current_page_num += 1
            _humanized_pause(page, current_page_num)

        while current_page_num <= target_pages:
            try:
                cards_data: list[dict] = page.evaluate(EXTRACT_CARDS_JS)

                if get_detail:
                    for idx, card in enumerate(cards_data):
                        detail = _extract_detail_page(page, context, idx)
                        card.update(detail)
                        page.wait_for_timeout(700)
                        try:
                            lbl_check = _get_page_label_text(page)
                            if str(current_page_num) not in lbl_check:
                                logger.warning(
                                    "Posición de página perdida tras detalle "
                                    "(esperado %d, etiqueta: '%s'). Renavegando...",
                                    current_page_num, lbl_check,
                                )
                                for _p in range(1, current_page_num):
                                    _lbl = _get_page_label_text(page)
                                    _go_to_next_page(page, _lbl)
                                    page.wait_for_timeout(800)
                        except Exception as _exc:
                            logger.debug("Verificación de página tras detalle: %s", _exc)
                else:
                    for card in cards_data:
                        card.update(_empty_detail())

                all_rows.extend(cards_data)
                pages_scraped += 1
            except Exception as exc:
                logger.warning("Fallo extrayendo página %d: %s", current_page_num, exc)
                failed_pages.append(current_page_num)

            if current_page_num >= target_pages:
                break

            try:
                label_text = _get_page_label_text(page)
                ok = _go_to_next_page(page, label_text)
                if not ok:
                    logger.error("Botón Sig. deshabilitado en página %d.", current_page_num)
                    stopped_early = True
                    error_detail  = "Botón Sig. deshabilitado"
                    break
            except Exception as exc:
                logger.error(
                    "Fallo avanzando desde página %d: %s — guardando lo recolectado.",
                    current_page_num, exc,
                )
                stopped_early = True
                error_detail  = str(exc)
                break

            current_page_num += 1
            _humanized_pause(page, current_page_num)

    except Exception as exc:
        logger.error("Fallo irrecuperable en scrape_servir_ofertas: %s", exc)
        stopped_early = True
        error_detail  = str(exc)
    finally:
        page.close()

    _write_excel(all_rows, output_path)

    result = {
        "output_file":           filename,
        "total_ofertas":         len(all_rows),
        "paginas_recorridas":    pages_scraped,
        "paginas_fallidas":      failed_pages,
        "recorrido_completo":    not stopped_early and not failed_pages and pages_scraped >= total_pages,
        "detalle_error":         error_detail,
    }
    logger.info("SERVIR: recorrido finalizado — %s", result)
    return result
